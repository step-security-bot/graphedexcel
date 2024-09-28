"""
This script extracts formulas from an Excel file and builds a dependency graph.
"""

from openpyxl import load_workbook
from collections import Counter
import networkx as nx
import re
import sys
from graph_visualizer import visualize_dependency_graph
from excel_parser import extract_references

# dictionary that stores the uniqe functions used in the formulas
# the key will be the funciton name and the value will be the number of times it was used
functions_dict = {}


def log(msg):
    """
    Log a message to the console if verbosity is enabled using the --verbose flag.
    """
    # if verbosity is enabled

    if "--verbose" in sys.argv:
        print(msg)


def stat_functions(cellvalue):
    """
    Extract the functions used in the formula and store them in a dictionary.
    This will be used to print the most used functions in the formulas.
    """

    # functions used in the formula
    cellfuncs = re.findall(r"[A-Z]+\(", cellvalue)
    log(f"  Functions used: {functions_dict}")
    for function in cellfuncs:
        # remove the "(" from the function name
        function = function[:-1]
        if function in functions_dict:
            functions_dict[function] += 1
        else:
            functions_dict[function] = 1


def extract_formulas_and_build_dependencies(file_path):
    """
    Extract formulas from an Excel file and build a dependency graph.
    """

    # Load the workbook
    wb = load_workbook(file_path, data_only=False)

    # Create a directed graph for dependencies
    graph = nx.DiGraph()

    # Iterate over all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log(f"-- Analyzing sheet: {sheet_name} --")

        # Iterate over all cells in the sheet and extract formulas
        for row in ws.iter_rows():
            for cell in row:
                # only interested in cells with formulas
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    # collect functions usage statistics
                    stat_functions(cell.value)

                    current_cell = f"{sheet_name}!{cell.coordinate}"
                    log(f"Formula in {current_cell}: {cell.value}")

                    graph.add_node(current_cell, sheet=sheet_name)

                    # Extract all referenced cells and ranges from the formula
                    direct_references, range_references, range_dependencies = (
                        extract_references(cell.value)
                    )

                    # all the referenced cells and cells from expanded ranges
                    # is added to the graph as nodes and edges
                    for ref_cell in direct_references:
                        if "!" not in ref_cell:
                            # No sheet specified, assume current sheet
                            refc = f"{sheet_name}!{ref_cell}"
                        else:
                            refc = ref_cell

                        log(f"  Depends on: {refc}")
                        # add the node
                        graph.add_node(refc, sheet=sheet_name)
                        # add the edge
                        graph.add_edge(current_cell, refc)

                    # If a range like A1:B3 is referenced, add the range definition as a node
                    for rng in range_references:
                        if "!" not in rng:
                            rng = f"{sheet_name}!{rng}"
                            range_sheet = sheet_name
                        else:
                            rng = rng
                            range_sheet = rng.split("!")[0]

                        graph.add_node(rng, sheet=sheet_name)
                        graph.add_edge(current_cell, rng)

                    # If a range like A1:B3 is referenced, add the
                    # edge between the cells within that range and
                    # the range istself
                    for single_cell, range_ref in range_dependencies.items():
                        if "!" not in range_ref:
                            range_ref = f"{sheet_name}!{range_ref}"
                            range_sheet = sheet_name
                        else:
                            range_ref = range_ref
                            range_sheet = range_ref.split("!")[0]

                        if "!" not in single_cell:
                            single_cell = f"{sheet_name}!{single_cell}"
                            cell_sheet = sheet_name
                        else:
                            single_cell = single_cell
                            cell_sheet = single_cell.split("!")[0]

                        # this is the single cell that points to the range it belongs to
                        graph.add_node(single_cell, sheet=cell_sheet)  # noqa
                        graph.add_node(range_ref, sheet=range_sheet)

                        # Then add the edge between the single cell and the range
                        graph.add_edge(range_ref, single_cell)
    return graph


def print_summary(graph, functionsdict):
    """
    Summarize a networkx DiGraph representing a dependency graph. And print the most used functions in the formulas
    """

    strpadsize = 28
    numpadsize = 5
    # 1. Print basic information about the graph

    print("=== Dependency Graph Summary ===")
    print(
        "Cell/Node count".ljust(strpadsize, " ")
        + str(graph.number_of_nodes()).rjust(numpadsize, " ")
    )
    print(
        "Dependency count".ljust(strpadsize, " ")
        + str(graph.number_of_edges()).rjust(numpadsize, " ")
    )
    print()

    # 2. Print the nodes with the highest degree
    degree_view = graph.degree()

    degree_counts = Counter(dict(degree_view))
    max_degree_node = degree_counts.most_common(10)
    print("=== Nodes with the highest degree ===")
    for node, degree in max_degree_node:
        print(f"{node.ljust(strpadsize)}{str(degree).rjust(numpadsize, ' ')} ")

    # 3. Print the most used functions
    print("\n=== Formula functions by count ===")
    sorted_functions = dict(
        sorted(functionsdict.items(), key=lambda item: item[1], reverse=True)
    )

    for function, count in sorted_functions.items():
        print(f"{function.ljust(strpadsize, ' ')}{str(count).rjust(numpadsize, ' ')}")


if __name__ == "__main__":
    path_to_excel = "Book1.xlsx"

    if len(sys.argv) > 1:
        path_to_excel = sys.argv[1]

    # Extract formulas and build the dependency graph
    dependency_graph = extract_formulas_and_build_dependencies(path_to_excel)

    print_summary(dependency_graph, functions_dict)

    # if --no-visualize is not passed as argument
    if "--no-visualize" not in sys.argv:
        print(
            "\033[1;30;40m\nVisualizing the graph of dependencies.\nThis might take a while...\033[0;37;40m\n"  # noqa
        )

        visualize_dependency_graph(dependency_graph, path_to_excel)
